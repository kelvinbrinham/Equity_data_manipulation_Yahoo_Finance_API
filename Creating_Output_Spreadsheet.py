import pandas as pd
import numpy as np
import requests as rq
import json
from os.path import exists
import yfinance as yf
from openpyxl import load_workbook
import openpyxl as xl
import datetime
from datetime import datetime as dt
from Retrieving_data import Ticker_list_stripped #Import the list of tickers


#Load the Universe spreadsheet
Universe_wb = load_workbook('Universe.xlsx')
#Add column for links to each ticker sheet in the universe workbook
Universe_wb['Sheet1'].cell(row=1, column=3).value = "Quarterly Earning dates (UTC)"

#Check if Ticker earning dates workbook exists, if it does not then there was no/insufficient data for that stock in Yahoo
i = 1
for Ticker in Ticker_list_stripped:
    i += 1
    Ticker_wb_exists = exists(f'Data/{Ticker}.xlsx')
    if not Ticker_wb_exists:
        Universe_wb['Sheet1'].cell(row=i, column=3).value = 'N/D' #Add No Data to relevant stocks
        continue

    #Load the Ticker earning date workbook
    Ticker_wb = load_workbook(f'Data/{Ticker}.xlsx')
    #Load the Ticker earning date worksheet
    Ticker_ws = Ticker_wb.active

    #Create new worksheet in universe workbook for individual ticker's earning dates
    Universe_wb.create_sheet(Ticker)
    Universe_wb_Ticker_ws = Universe_wb[Ticker]

    #Add the ticker earnings worksheet to the universe workbook
    #I do this value by value because there is no reliable way to simply copy a worksheet from one workbook to another
    for row in Ticker_ws:
        for cell in row:
            Universe_wb_Ticker_ws[cell.coordinate].value = cell.value

    #Link to the output universe workbook ticker worksheeta
    link = f'Universe_OUTPUT.xlsx#{Ticker}!A1'

    #Add links to each ticker worksheet within the universe workbook
    Universe_wb['Sheet1'].cell(row=i, column=3).hyperlink = link
    Universe_wb['Sheet1'].cell(row=i, column=3).value = Ticker
    Universe_wb['Sheet1'].cell(row=i, column=3).style = "Hyperlink"

#Save the output universe workbook
Universe_wb.save('Universe_OUTPUT.xlsx')
